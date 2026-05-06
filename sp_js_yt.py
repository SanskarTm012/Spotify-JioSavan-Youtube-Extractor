# Import libraries
import re
import sys
import time
import json
import random
import logging
import argparse
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import re

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ── Windows-safe logging ──────────────────────────────────────────────────────
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("extractor.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
HEADERS_POOL = [
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    },
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
        "Accept-Language": "en-GB,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    },
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    },
]

# ── Faster Delays ─────────────────────────────
MIN_DELAY = 0.5
MAX_DELAY = 1.5
RETRY_DELAYS = [5, 10, 20]


# ── Helpers ───────────────────────────────────────────────────────────────────

def random_headers():
    return random.choice(HEADERS_POOL).copy()


def polite_sleep(extra: float = 0.0):
    t = random.uniform(MIN_DELAY, MAX_DELAY) + extra
    time.sleep(t)


def get_with_retry(url: str, session: requests.Session, max_retries: int = 3,
                   extra_headers: dict = None, **kwargs):
    """GET with exponential back-off on 429/5xx."""
    headers = random_headers()
    if extra_headers:
        headers.update(extra_headers)
    for attempt, wait in enumerate(RETRY_DELAYS[:max_retries], start=1):
        try:
            resp = session.get(url, headers=headers, timeout=25, **kwargs)
            if resp.status_code == 429:
                log.warning(f"  Rate limited (attempt {attempt}). Waiting {wait}s...")
                time.sleep(wait)
                continue
            if resp.status_code >= 500:
                log.warning(f"  Server error {resp.status_code} (attempt {attempt}). Waiting {wait}s...")
                time.sleep(wait)
                continue
            return resp
        except requests.RequestException as e:
            log.warning(f"  Request error (attempt {attempt}): {e}. Waiting {wait}s...")
            time.sleep(wait)
    return None


def normalize_count(text: str) -> str:
    if not text:
        return "N/A"
    cleaned = text.strip()
    for ch in [",", "\u202f", "\xa0", " "]:
        cleaned = cleaned.replace(ch, "")
    return cleaned if cleaned else "N/A"


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


# ── Platform detector ─────────────────────────────────────────────────────────

def detect_platform(url: str) -> str:
    u = url.lower().strip()
    if "spotify.com" in u:
        return "spotify"
    if "jiosaavn.com" in u or "saavn.com" in u:
        return "jiosaavn"
    if "youtube.com" in u or "youtu.be" in u:
        return "youtube"
    return "unknown"


# ── Spotify (Selenium-based - proven working) ─────────────────────────────────

def _make_spotify_driver():
    """Create a headless Chrome driver with anti-detection settings."""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
    )
    return driver


def extract_spotify(url: str, driver) -> dict:
    result = {
        "spotify_title": "N/A",
        "spotify_streams": "N/A",
        "spotify_status": "ok",
    }

    tid_match = re.search(r"spotify\.com/(?:intl-[a-z]+/)?track/([A-Za-z0-9]+)", url)
    if not tid_match:
        result["spotify_status"] = "invalid_url"
        return result

    clean_url = f"https://open.spotify.com/track/{tid_match.group(1)}"

    try:
        driver.get(clean_url)
        wait = WebDriverWait(driver, 10)

        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//section[@data-testid='track-page']")
        ))

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        # Title
        try:
            el = driver.find_element(By.CSS_SELECTOR, "[data-testid='entityTitle'] h1")
            result["spotify_title"] = el.text.strip()
        except:
            pass

        # Streams
        try:
            pc_el = driver.find_element(By.CSS_SELECTOR, "[data-testid='playcount']")
            result["spotify_streams"] = pc_el.text.replace(",", "")
        except:
            pass

        if result["spotify_streams"] == "N/A":
            result["spotify_status"] = "streams_not_found"

    except Exception as e:
        result["spotify_status"] = f"error: {e}"

    return result

# ── JioSaavn ──────────────────────────────────────────────────────────────────

def extract_jiosaavn(url: str, session: requests.Session) -> dict:
    result = {
        "jiosaavn_title": "N/A",
        "jiosaavn_streams": "N/A",
        "jiosaavn_status": "ok",
    }

    # -- Method 1: JioSaavn internal API --
    try:
        api_url = (
            f"https://www.jiosaavn.com/api.php?__call=webapi.get&token=&type=song"
            f"&p=1&n=1&url={url}&ctx=web6dot0&_format=json&_marker=0"
        )
        headers = random_headers()
        headers.update({
            "Referer": "https://www.jiosaavn.com/",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
        })
        resp = session.get(api_url, headers=headers, timeout=20)
        if resp and resp.status_code == 200:
            raw = resp.text.strip()
            m = re.search(r"^\w+\((.*)\)$", raw, re.DOTALL)
            if m:
                raw = m.group(1)
            data = json.loads(raw)
            songs = data.get("songs", [data]) if "songs" in data else [data]
            if songs:
                song = songs[0]
                title = song.get("title") or song.get("song", "")
                import html as html_mod
                title = html_mod.unescape(title)
                play_count = song.get("play_count", "") or song.get("playCount", "")
                if title:
                    result["jiosaavn_title"] = title
                if play_count:
                    result["jiosaavn_streams"] = str(play_count)
                    return result
    except Exception as e:
        log.warning(f"  JioSaavn API failed: {e}")

    # -- Method 2: Scrape the page HTML --
    try:
        resp = get_with_retry(url, session)
        if resp is None or resp.status_code != 200:
            result["jiosaavn_status"] = f"HTTP_{resp.status_code if resp else 'no_response'}"
            return result

        html = resp.text
        soup = BeautifulSoup(html, "lxml")

        og_title = soup.find("meta", property="og:title")
        if og_title and result["jiosaavn_title"] == "N/A":
            result["jiosaavn_title"] = og_title.get("content", "N/A")

        for pat in [
            r'"play_count"\s*:\s*"?(\d+)"?',
            r'"playCount"\s*:\s*"?(\d+)"?',
            r'"stream_count"\s*:\s*"?(\d+)"?',
            r'play.count["\s:]+(\d+)',
        ]:
            m = re.search(pat, html, re.IGNORECASE | re.DOTALL)
            if m:
                result["jiosaavn_streams"] = m.group(1)
                return result

        # JSON-LD fallback
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, dict):
                    count = data.get("interactionStatistic", {}).get("userInteractionCount")
                    if count:
                        result["jiosaavn_streams"] = str(count)
                        return result
            except Exception:
                pass

        result["jiosaavn_status"] = "streams_not_found"
    except Exception as e:
        result["jiosaavn_status"] = f"error: {e}"

    return result


# ── YouTube ───────────────────────────────────────────────────────────────────
ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def clean_excel_text(value):
    if value is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))

def get_video_id(url: str):
    match = re.search(r"(?:v=|youtu\.be/)([a-zA-Z0-9_-]{11})", url)
    return match.group(1) if match else None

def extract_youtube(url: str) -> dict:
    API_KEY = "AIzaSyCksPVmnlCoFILqjXVdFGKfzW6pmC9b4oQ"

    result = {
        "yt_title": "N/A",
        "yt_views": "N/A",
        "yt_likes": "N/A",
        "yt_status": "ok",
    }

    video_id = get_video_id(url)
    if not video_id:
        result["yt_status"] = "invalid_url"
        return result

    api_url = (
        f"https://www.googleapis.com/youtube/v3/videos"
        f"?part=snippet,statistics&id={video_id}&key={API_KEY}"
    )

    try:
        resp = requests.get(api_url, timeout=15)

        if resp.status_code != 200:
            result["yt_status"] = f"HTTP_{resp.status_code}"
            return result

        data = resp.json()

        if not data.get("items"):
            result["yt_status"] = "no_data"
            return result

        item = data["items"][0]

        # Title
        result["yt_title"] = item["snippet"].get("title", "N/A")

        stats = item.get("statistics", {})

        # EXACT VALUES
        result["yt_views"] = stats.get("viewCount", "N/A")
        result["yt_likes"] = stats.get("likeCount", "N/A")

    except Exception as e:
        result["yt_status"] = f"error: {e}"

    return result

# def extract_youtube(url: str) -> dict:
#     result = {
#         "yt_title": "N/A",
#         "yt_views": "N/A",
#         "yt_likes": "N/A",
#         "yt_status": "ok",
#     }

#     options = Options()
#     options.add_argument("--headless=new")
#     options.add_argument("--disable-blink-features=AutomationControlled")

#     driver = webdriver.Chrome(options=options)

#     try:
#         driver.get(url)
#         wait = WebDriverWait(driver, 15)

#         # Wait for title
#         wait.until(EC.presence_of_element_located((By.TAG_NAME, "h1")))

#         time.sleep(3)

#         # Title
#         try:
#             title_el = WebDriverWait(driver, 15).until(
#                 EC.presence_of_element_located(
#                     (By.CSS_SELECTOR, "ytd-watch-metadata h1 yt-formatted-string")
#                 )
#             )

#             title = title_el.get_attribute("title") or title_el.text
#             if title:
#                 result["yt_title"] = title.strip()

#         except Exception:
#             pass

#         # --- YouTube Views ---
#         try:
#             # Method 1: meta tag (MOST RELIABLE)
#             meta_views = driver.find_element(By.XPATH, "//meta[@itemprop='interactionCount']")
#             if meta_views:
#                 result["yt_views"] = meta_views.get_attribute("content")
#         except Exception:
#             pass

#         # Fallback (if meta fails)
#         if result["yt_views"] == "N/A":
#             try:
#                 scripts = driver.find_elements(By.XPATH, "//script[contains(text(),'viewCount')]")
#                 for script in scripts:
#                     text = script.get_attribute("innerHTML")

#                     match = re.search(r'"viewCount":"(\d+)"', text)
#                     if match:
#                         result["yt_views"] = match.group(1)
#                         break
#             except Exception:
#                 pass

#         try:
#             like_button = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'like')]")
#             like_text = like_button.get_attribute("aria-label")

#             # Extract number using regex
#             match = re.search(r'([\d,]+)', like_text)
#             if match:
#                 result["yt_likes"] = match.group(1).replace(",", "")
#         except:
#             pass

#     except Exception as e:
#         result["yt_status"] = f"error: {e}"

#     finally:
#         driver.quit()

#     return result

# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADERS = [
    "#",
    "Spotify Link", "Spotify Title", "Spotify Streams", "Spotify Status",
    "JioSaavn Link", "JioSaavn Title", "JioSaavn Streams", "JioSaavn Status",
    "YouTube Link", "YouTube Title", "YouTube Views", "YouTube Likes", "YouTube Status",
    "Extracted At",
]
COL_WIDTHS = [4, 45, 35, 15, 20, 45, 35, 15, 20, 45, 35, 15, 15, 20, 22]

HEADER_FILL = PatternFill("solid", start_color="1DB954")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT   = Font(name="Arial", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

STATUS_OK_FILL   = PatternFill("solid", start_color="EAF7ED")
STATUS_ERR_FILL  = PatternFill("solid", start_color="FDE8E8")
STATUS_WARN_FILL = PatternFill("solid", start_color="FFF8E1")


def setup_output_wb(path: Path):
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        start_row = ws.max_row + 1
        log.info(f"Resuming output file at row {start_row}.")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20
        for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
        start_row = 2
    return wb, ws, start_row


def row_fill(sp_status, js_status, yt_status):
    statuses = [sp_status, js_status, yt_status]
    if all(s == "ok" for s in statuses):
        return STATUS_OK_FILL
    if all(s not in ("ok",) for s in statuses):
        return STATUS_ERR_FILL
    return STATUS_WARN_FILL


def write_row(ws, row_num: int, serial: int,
              sp_url, sp_data: dict,
              js_url, js_data: dict,
              yt_url, yt_data: dict):

    fill = row_fill(
        sp_data["spotify_status"],
        js_data["jiosaavn_status"],
        yt_data["yt_status"],
    )
    values = [
        serial,
        safe_str(sp_url),
        sp_data["spotify_title"],
        sp_data["spotify_streams"],
        sp_data["spotify_status"],
        safe_str(js_url),
        js_data["jiosaavn_title"],
        js_data["jiosaavn_streams"],
        js_data["jiosaavn_status"],
        safe_str(yt_url),
        yt_data["yt_title"],
        yt_data["yt_views"],
        yt_data["yt_likes"],
        yt_data["yt_status"],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    left_cols = {2, 3, 6, 7, 10, 11}
    for col_idx, value in enumerate(values, 1):
        safe_value = clean_excel_text(value)
        cell = ws.cell(row=row_num, column=col_idx, value=safe_value)
        cell.font = CELL_FONT
        cell.fill = fill
        cell.alignment = LEFT if col_idx in left_cols else CENTER


# ── Main ──────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str,
         sp_col: str = "Spotify Link",
         js_col: str = "JioSaavn Link",
         yt_col: str = "YouTube Links",
         start_from: int = 1,
         max_rows: int = None):

    input_file  = Path(input_path)
    output_file = Path(output_path)

    if not input_file.exists():
        log.error(f"Input file not found: {input_file}")
        sys.exit(1)

    df = pd.read_excel(input_file, dtype=str)
    df.dropna(how="all", inplace=True)

    log.info(f"Columns found in input: {list(df.columns)}")

    # Validate columns
    for col_name, col_label in [(sp_col, "Spotify"), (js_col, "JioSaavn"), (yt_col, "YouTube")]:
        if col_name not in df.columns:
            log.error(f"{col_label} column '{col_name}' not found. Available: {list(df.columns)}")
            sys.exit(1)

    total = len(df)
    log.info(f"Total rows: {total}")

    if max_rows:
        df = df.head(max_rows)
    df = df.iloc[start_from - 1:]

    wb, ws, excel_row = setup_output_wb(output_file)
    session = requests.Session()
    session.headers.update(random_headers())

    ok_count = fail_count = partial_count = 0
    # CREATE DRIVER ONCE
    spotify_driver = _make_spotify_driver()

    for idx, (_, row) in enumerate(df.iterrows(), start=start_from):
        sp_url = safe_str(row.get(sp_col, ""))
        js_url = safe_str(row.get(js_col, ""))
        yt_url = safe_str(row.get(yt_col, ""))

        log.info(f"[{idx}/{total}] Processing row {idx}...")

        # --- Spotify ---
        if sp_url and sp_url.lower() not in ("nan", ""):
            log.info(f"  Spotify: {sp_url[:60]}")
            sp_data = extract_spotify(sp_url, spotify_driver)
            polite_sleep()
        else:
            sp_data = {"spotify_title": "N/A", "spotify_streams": "N/A", "spotify_status": "no_url"}

        # --- JioSaavn ---
        if js_url and js_url.lower() not in ("nan", ""):
            log.info(f"  JioSaavn: {js_url[:60]}")
            js_data = extract_jiosaavn(js_url, session)
            polite_sleep()
        else:
            js_data = {"jiosaavn_title": "N/A", "jiosaavn_streams": "N/A", "jiosaavn_status": "no_url"}

        # --- YouTube ---
        if yt_url and yt_url.lower() not in ("nan", ""):
            log.info(f"  YouTube: {yt_url[:60]}")
            yt_data = extract_youtube(yt_url)
            polite_sleep()
        else:
            yt_data = {"yt_title": "N/A", "yt_views": "N/A", "yt_likes": "N/A", "yt_status": "no_url"}

        # Log result summary
        log.info(
            f"  => SP streams={sp_data['spotify_streams']} | "
            f"JS streams={js_data['jiosaavn_streams']} | "
            f"YT views={yt_data['yt_views']} likes={yt_data['yt_likes']}"
        )

        # Track counts
        statuses = [sp_data["spotify_status"], js_data["jiosaavn_status"], yt_data["yt_status"]]
        if all(s == "ok" for s in statuses):
            ok_count += 1
        elif all(s not in ("ok",) for s in statuses):
            fail_count += 1
        else:
            partial_count += 1

        # Write + save immediately (crash-safe)
        write_row(ws, excel_row, idx, sp_url, sp_data, js_url, js_data, yt_url, yt_data)
        wb.save(output_file)
        excel_row += 1

        # Rotate UA every 50 rows
        if idx % 50 == 0:
            session.headers.update(random_headers())
            log.info("  Rotated User-Agent.")

        # Extra pause every 100 rows
        if idx % 100 == 0:
            log.info("  Taking a longer break (100-row checkpoint)...")
            time.sleep(random.uniform(8, 15))

    log.info(
        f"\nDone! {ok_count} fully OK | {partial_count} partial | {fail_count} failed."
    )
    log.info(f"Output saved to: {output_file.resolve()}")
    spotify_driver.quit()

# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Music Data Extractor v3 - Spotify (Selenium) / JioSaavn / YouTube"
    )
    parser.add_argument("input",  help="Input Excel file (.xlsx)")
    parser.add_argument("output", help="Output Excel file (.xlsx)")
    parser.add_argument("--sp-col", default="Spotify Link",   help="Spotify column name")
    parser.add_argument("--js-col", default="JioSaavn Link",  help="JioSaavn column name")
    parser.add_argument("--yt-col", default="YouTube Links",  help="YouTube column name")
    parser.add_argument("--start",  type=int, default=1,      help="Resume from row number")
    parser.add_argument("--max",    type=int, default=None,   help="Max rows to process")
    args = parser.parse_args()

    main(
        input_path=args.input,
        output_path=args.output,
        sp_col=args.sp_col,
        js_col=args.js_col,
        yt_col=args.yt_col,
        start_from=args.start,
        max_rows=args.max,
    )